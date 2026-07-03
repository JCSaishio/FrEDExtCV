"""
FrED Fiber Measure (with Streaming v3 - WiFi)
=============================================

Real-time fiber diameter measurement from a USB camera feed, with the ability
to **stream the measured diameter to a FrED Raspberry Pi over WiFi (TCP)**.

The Pi runs as a self-contained WiFi hotspot (see ``setup_hotspot.sh`` in the
``fred-device-external-cv-pi4v3`` folder). This laptop joins that hotspot and
connects to the Pi as a TCP client, then streams each measurement to it.

Features
--------
* Live border (edge) detection of a bright fiber on a dark background.
* Robust diameter measurement that tolerates a tilted fiber and captures
  diameter variation along the fiber (median / min / max / std).
* Pixel <-> real-world calibration so measurements are reported in real units
  (mm, um, etc.). Calibration is persisted to ``calibration.json``.
* Start / pause experiment recording at will. Each recorded frame becomes one
  row of data. On pause, the user is asked to confirm before a timestamped
  ``.csv`` file is written into the ``Data`` folder next to this script.
* **Streaming**: connect to a serial (USB) port and stream each measurement to
  the FrED Pi as newline-delimited JSON. The Pi reads it via
  ``external_diameter.py`` and graphs it in place of its old camera feed.

Streaming wire protocol (newline-delimited JSON, one object per line over TCP)::

    {"v": 1, "d": 0.352, "u": "mm", "t": 12.345, "found": true}

Run with:  python fiber_measure.py
"""

import base64
import csv
import datetime as _dt
import io
import json
import os
import queue
import socket
import threading
import time
import tkinter as tk
from tkinter import filedialog, font as tkfont, messagebox, simpledialog, ttk

import cv2
import numpy as np
from PIL import Image, ImageTk

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "Data")
CALIB_FILE = os.path.join(BASE_DIR, "calibration.json")

# --------------------------------------------------------------------------- #
# FrED Pi hotspot defaults (must match external_diameter.py / setup_hotspot.sh)
# --------------------------------------------------------------------------- #
PI_HOTSPOT_SSID = "FrED_Pi"
PI_HOTSPOT_PASSWORD = "fredfiber123"
PI_DEFAULT_IP = "192.168.4.1"   # the Pi's address while acting as the hotspot
PI_DEFAULT_PORT = 5005          # TCP port the Pi listens on

# Column order shared by the CSV and Excel exports.
EXPORT_FIELDS = [
    "timestamp", "elapsed_s", "frame", "diameter_px", "diameter_real",
    "units", "min_px", "max_px", "std_px", "length_px", "angle_deg",
]


# --------------------------------------------------------------------------- #
# Calibration
# --------------------------------------------------------------------------- #
class Calibration:
    """Stores and persists the pixel -> real-world conversion factor."""

    def __init__(self):
        # units_per_pixel: how many real-world units one pixel represents.
        self.units_per_pixel = None  # None -> not calibrated (report in px)
        self.units = "mm"
        self.load()

    @property
    def is_calibrated(self):
        return self.units_per_pixel is not None and self.units_per_pixel > 0

    def to_real(self, value_px):
        """Convert a pixel measurement to real units (or return px if not set)."""
        if self.is_calibrated:
            return value_px * self.units_per_pixel
        return value_px

    def set_from_reference(self, measured_px, known_real, units):
        if measured_px <= 0:
            raise ValueError("Measured pixel size must be > 0.")
        self.units_per_pixel = float(known_real) / float(measured_px)
        self.units = units
        self.save()

    def set_factor(self, units_per_pixel, units):
        self.units_per_pixel = float(units_per_pixel)
        self.units = units
        self.save()

    def clear(self):
        self.units_per_pixel = None
        self.save()

    def load(self):
        try:
            with open(CALIB_FILE, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            self.units_per_pixel = data.get("units_per_pixel")
            self.units = data.get("units", "mm")
        except (FileNotFoundError, json.JSONDecodeError, ValueError):
            pass

    def save(self):
        data = {"units_per_pixel": self.units_per_pixel, "units": self.units}
        try:
            with open(CALIB_FILE, "w", encoding="utf-8") as fh:
                json.dump(data, fh, indent=2)
        except OSError:
            pass


# --------------------------------------------------------------------------- #
# WiFi (TCP) streaming to the FrED Raspberry Pi
# --------------------------------------------------------------------------- #
class TCPStreamer:
    """Send diameter measurements to the FrED Pi over a WiFi TCP socket.

    The Pi (``external_diameter.py``) is the server: it runs as a hotspot and
    listens on a port. This class connects to it as a client and sends each
    measurement as one line of newline-delimited JSON::

        {"v": 1, "d": 0.352, "u": "mm", "t": 12.345, "found": true}
    """

    PROTOCOL_VERSION = 1
    CONNECT_TIMEOUT = 5.0   # seconds to wait when opening the connection

    def __init__(self):
        self.sock = None
        self.host = None
        self.port = None
        # Messages received FROM the Pi (status / experiment data) are parsed on
        # a background thread and queued for the Tk main loop to handle safely.
        self.rx_queue = queue.Queue()
        self._reader = None
        self._stop = threading.Event()

    @property
    def is_open(self):
        return self.sock is not None

    def open(self, host, port):
        """Connect to the Pi at host:port; raises on failure."""
        self.close()
        sock = socket.create_connection((host, int(port)),
                                        timeout=self.CONNECT_TIMEOUT)
        sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_NODELAY, 1)
        self.sock = sock
        self.host = host
        self.port = int(port)
        # Start the background reader for Pi -> laptop messages.
        self._stop.clear()
        self._reader = threading.Thread(target=self._reader_loop, daemon=True)
        self._reader.start()

    def close(self):
        self._stop.set()
        if self.sock is not None:
            try:
                self.sock.close()
            except Exception:
                pass
        self.sock = None

    @property
    def endpoint(self):
        if self.host is None:
            return "?"
        return f"{self.host}:{self.port}"

    def _reader_loop(self):
        """Read newline-delimited JSON from the Pi and queue each message."""
        buffer = b""
        sock = self.sock
        if sock is None:
            return
        sock.settimeout(0.5)
        while not self._stop.is_set():
            try:
                data = sock.recv(4096)
            except socket.timeout:
                continue
            except Exception:
                break
            if not data:        # Pi closed the connection
                break
            buffer += data
            while b"\n" in buffer:
                line, buffer = buffer.split(b"\n", 1)
                text = line.decode("utf-8", errors="ignore").strip()
                if not text:
                    continue
                try:
                    self.rx_queue.put(json.loads(text))
                except (ValueError, json.JSONDecodeError):
                    pass

    def _send_obj(self, obj):
        """Send one JSON object; closes the link on error. Returns success."""
        if not self.is_open:
            return False
        try:
            self.sock.sendall((json.dumps(obj) + "\n").encode("utf-8"))
            return True
        except Exception:
            self.close()
            return False

    def send(self, diameter, units, elapsed, found):
        """Send one diameter measurement."""
        return self._send_obj({
            "v": self.PROTOCOL_VERSION,
            "d": round(float(diameter), 5),
            "u": units,
            "t": round(float(elapsed), 3),
            "found": bool(found),
        })

    def send_command(self, obj):
        """Send an experiment command (experiment / get_data / abort)."""
        return self._send_obj(obj)


# --------------------------------------------------------------------------- #
# Fiber detection / measurement
# --------------------------------------------------------------------------- #
class FiberDetector:
    """
    Detects a bright, elongated fiber on a dark background and measures its
    diameter (the short dimension), tolerating tilt.

    Pipeline:
        grayscale -> Gaussian blur -> threshold (Otsu or manual) ->
        morphological clean-up -> largest contour -> minAreaRect for
        orientation -> rotate mask flat -> per-column thickness profile.
    """

    # Preset detection parameters (used at start-up and by "Reset to defaults").
    DEFAULTS = {
        "use_otsu": True,       # auto threshold
        "manual_thresh": 110,   # used when use_otsu is False
        "blur_ksize": 5,        # odd kernel size for Gaussian blur
        "min_area": 500,        # ignore blobs smaller than this (px^2)
    }

    def __init__(self):
        self.use_otsu = FiberDetector.DEFAULTS["use_otsu"]
        self.manual_thresh = FiberDetector.DEFAULTS["manual_thresh"]
        self.blur_ksize = FiberDetector.DEFAULTS["blur_ksize"]
        self.min_area = FiberDetector.DEFAULTS["min_area"]

    def _threshold(self, gray):
        k = max(1, self.blur_ksize)
        if k % 2 == 0:
            k += 1
        blur = cv2.GaussianBlur(gray, (k, k), 0)
        if self.use_otsu:
            _, mask = cv2.threshold(
                blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
            )
        else:
            _, mask = cv2.threshold(
                blur, self.manual_thresh, 255, cv2.THRESH_BINARY
            )
        # Clean up: remove specks, then fill small gaps inside the fiber.
        kernel = np.ones((3, 3), np.uint8)
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=1)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)
        return mask

    def process(self, frame):
        """
        Returns (annotated_bgr_frame, mask, result_dict).

        ``mask`` is the binary detection image (after blur/threshold/morphology)
        with the detected contour and diameter drawn on it, so the processed
        feed shows exactly where edge detection happens and how the parameters
        affect it.

        result_dict keys:
            found (bool), diameter_px, min_px, max_px, std_px, length_px,
            angle_deg, n_samples
        """
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        mask = self._threshold(gray)

        contours, _ = cv2.findContours(
            mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
        )
        annotated = frame.copy()
        # Colour copy of the mask so we can draw coloured overlays on it.
        mask_view = cv2.cvtColor(mask, cv2.COLOR_GRAY2BGR)
        result = {
            "found": False,
            "diameter_px": 0.0,
            "min_px": 0.0,
            "max_px": 0.0,
            "std_px": 0.0,
            "length_px": 0.0,
            "angle_deg": 0.0,
            "n_samples": 0,
        }

        if not contours:
            return annotated, mask_view, result

        largest = max(contours, key=cv2.contourArea)
        if cv2.contourArea(largest) < self.min_area:
            # Show the rejected blob outline in red so the Min-area effect is visible.
            cv2.drawContours(mask_view, [largest], 0, (0, 0, 255), 2)
            return annotated, mask_view, result

        rect = cv2.minAreaRect(largest)  # ((cx,cy),(w,h),angle)
        (cx, cy), (rw, rh), angle = rect

        # Orient so the longer side (fiber length) becomes horizontal.
        if rw < rh:
            rot_angle = angle + 90.0
            length_guess = rh
        else:
            rot_angle = angle
            length_guess = rw

        profile = self._thickness_profile(mask, (cx, cy), rot_angle)

        if profile.size == 0:
            return annotated, mask_view, result

        diameter = float(np.median(profile))
        result.update(
            found=True,
            diameter_px=diameter,
            min_px=float(np.min(profile)),
            max_px=float(np.max(profile)),
            std_px=float(np.std(profile)),
            length_px=float(length_guess),
            angle_deg=float(rot_angle),
            n_samples=int(profile.size),
        )

        # ---- overlay (drawn on both the live frame and the mask view) ---- #
        box = cv2.boxPoints(rect).astype(np.int32)
        theta = np.deg2rad(rot_angle)
        px, py = -np.sin(theta), np.cos(theta)  # perpendicular direction
        half = diameter / 2.0
        p1 = (int(cx - px * half), int(cy - py * half))
        p2 = (int(cx + px * half), int(cy + py * half))
        for canvas in (annotated, mask_view):
            cv2.drawContours(canvas, [box], 0, (0, 255, 0), 2)
            # Short perpendicular tick at the fiber centre showing the diameter.
            cv2.line(canvas, p1, p2, (0, 0, 255), 2)
            cv2.circle(canvas, (int(cx), int(cy)), 3, (0, 0, 255), -1)

        return annotated, mask_view, result

    @staticmethod
    def _thickness_profile(mask, center, rot_angle):
        """
        Rotate the mask so the fiber is horizontal, then count white pixels in
        each column to build a thickness (diameter) profile in pixels.
        """
        h, w = mask.shape[:2]
        M = cv2.getRotationMatrix2D(center, rot_angle, 1.0)
        rotated = cv2.warpAffine(
            mask, M, (w, h), flags=cv2.INTER_NEAREST, borderValue=0
        )
        col_counts = (rotated > 0).sum(axis=0).astype(np.float32)

        valid = col_counts[col_counts > 0]
        if valid.size == 0:
            return np.array([], dtype=np.float32)

        # Trim the outer 10% of the fiber span to avoid tapered ends.
        nz = np.flatnonzero(col_counts > 0)
        x0, x1 = nz[0], nz[-1]
        span = x1 - x0
        if span > 20:
            margin = int(span * 0.10)
            x0 += margin
            x1 -= margin
        profile = col_counts[x0 : x1 + 1]
        profile = profile[profile > 0]
        return profile


# --------------------------------------------------------------------------- #
# GUI application
# --------------------------------------------------------------------------- #
class FiberApp:
    def __init__(self, root):
        self.root = root
        self.root.title("FrED Fiber Measure - WiFi Streaming (v3)")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.minsize(960, 620)
        # F11 toggles true fullscreen, Escape leaves it.
        self._fullscreen = False
        self.root.bind("<F11>", self._toggle_fullscreen)
        self.root.bind("<Escape>", self._exit_fullscreen)

        self.calib = Calibration()
        self.detector = FiberDetector()

        self.cap = None
        self.camera_index = 0
        self.last_result = None

        # Size each video feed is scaled to (updated as the window resizes).
        self._feed_w = 480
        self._feed_h = 360

        # Experiment recording state
        self.recording = False
        self.records = []          # list of dicts (one per recorded frame)
        self.exp_start_time = None
        self.frame_counter = 0
        self.save_dir = DATA_DIR   # where Pause & Save writes (default: Data/)

        # Streaming-to-Pi state
        self.streamer = TCPStreamer()
        self.streaming = False
        self.stream_start_time = None
        self._last_stream_t = 0.0
        self.STREAM_INTERVAL = 0.05   # seconds between sent samples (~20 Hz)

        # Remote-experiment state
        self.exp_save_dir = DATA_DIR
        self.exp_data_ready = False

        self._build_ui()
        self._set_initial_geometry()
        self._update_calib_label()
        self.open_camera()
        self._loop()

    def _set_initial_geometry(self):
        """Open at a size where everything fits with a little free space."""
        self.root.update_idletasks()
        bbox = self.panel_canvas.bbox("all")
        panel_w = (bbox[2] - bbox[0]) if bbox else 480
        panel_h = (bbox[3] - bbox[1]) if bbox else 800

        # Width: control panel + a comfortable video column; height: tall enough
        # for the whole panel plus the status bar and some breathing room.
        want_w = panel_w + 760 + 40
        want_h = panel_h + 70

        # Never exceed the available screen.
        max_w = int(self.root.winfo_screenwidth() * 0.96)
        max_h = int(self.root.winfo_screenheight() * 0.95)
        win_w, win_h = min(want_w, max_w), min(want_h, max_h)

        # Centre the window on screen.
        x = max(0, (self.root.winfo_screenwidth() - win_w) // 2)
        y = max(0, (self.root.winfo_screenheight() - win_h) // 3)
        self.root.geometry(f"{win_w}x{win_h}+{x}+{y}")

    # ------------------------------------------------------------------ #
    # UI construction
    # ------------------------------------------------------------------ #
    def _apply_style(self):
        """Enlarge the fonts/widgets so the control panel is bigger and clearer."""
        # Bump the named fonts every ttk widget inherits from (works on all themes).
        for name, size in (("TkDefaultFont", 11), ("TkTextFont", 11),
                           ("TkMenuFont", 11)):
            try:
                tkfont.nametofont(name).configure(size=size)
            except tk.TclError:
                pass
        style = ttk.Style()
        style.configure(".", font=("Segoe UI", 11))
        style.configure("TButton", font=("Segoe UI", 11), padding=5)
        style.configure("TCheckbutton", font=("Segoe UI", 11))
        style.configure("TLabelframe.Label", font=("Segoe UI", 12, "bold"))
        # Larger font for the port dropdown list.
        self.root.option_add("*TCombobox*Listbox.font", ("Segoe UI", 11))

    def _build_ui(self):
        self._apply_style()
        main = ttk.Frame(self.root, padding=(8, 5))
        main.pack(fill=tk.BOTH, expand=True)
        # Video area expands with the window; the control column keeps its width.
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=0)
        main.rowconfigure(0, weight=1)

        # --- Video area: live feed on top, processed feed below -------------- #
        video_area = ttk.Frame(main)
        video_area.grid(row=0, column=0, sticky="nsew")
        video_area.columnconfigure(0, weight=1)
        video_area.rowconfigure(1, weight=1)   # live feed
        video_area.rowconfigure(3, weight=1)   # processed feed
        video_area.bind("<Configure>", self._on_video_resize)

        ttk.Label(video_area, text="Live (detected)",
                  font=("Segoe UI", 11, "bold")).grid(row=0, column=0, pady=(0, 2))
        self.video_label = ttk.Label(video_area, anchor="center")
        self.video_label.grid(row=1, column=0, sticky="nsew", padx=4, pady=2)
        ttk.Label(video_area, text="Processed (mask / edge detection)",
                  font=("Segoe UI", 11, "bold")).grid(row=2, column=0, pady=(4, 2))
        self.proc_label = ttk.Label(video_area, anchor="center")
        self.proc_label.grid(row=3, column=0, sticky="nsew", padx=4, pady=2)

        # --- Right side: a notebook with the controls and the experiment tab - #
        self.notebook = ttk.Notebook(main)
        self.notebook.grid(row=0, column=1, sticky="ns")
        tab_measure = ttk.Frame(self.notebook)
        self.notebook.add(tab_measure, text="Measure & Stream")
        self._experiment_tab = ttk.Frame(self.notebook)
        self.notebook.add(self._experiment_tab, text="Experiment (FrED)")

        # --- Control panel: sizes to its content; scrollbar only when needed - #
        panel_container = ttk.Frame(tab_measure)
        panel_container.pack(fill=tk.BOTH, expand=True)
        panel_container.rowconfigure(0, weight=1)
        panel_container.columnconfigure(0, weight=1)

        self.panel_canvas = tk.Canvas(panel_container, borderwidth=0,
                                      highlightthickness=0)
        self.panel_vsb = ttk.Scrollbar(panel_container, orient="vertical",
                                       command=self.panel_canvas.yview)
        self.panel_canvas.configure(yscrollcommand=self.panel_vsb.set)
        self.panel_canvas.grid(row=0, column=0, sticky="nsew")
        self.panel_vsb.grid(row=0, column=1, sticky="ns")
        self.panel_vsb.grid_remove()   # hidden until the content is taller than the view

        panel = ttk.Frame(self.panel_canvas, padding=(12, 0))
        self.panel_canvas.create_window((0, 0), window=panel, anchor="nw")
        panel.bind("<Configure>", lambda e: self._update_panel_scroll())
        self.panel_canvas.bind("<Configure>", lambda e: self._update_panel_scroll())

        def _on_wheel(event):
            if self.panel_vsb.winfo_ismapped():   # only scroll when scrollbar is active
                self.panel_canvas.yview_scroll(int(-event.delta / 120), "units")
        self.panel_canvas.bind(
            "<Enter>", lambda e: self.panel_canvas.bind_all("<MouseWheel>", _on_wheel))
        self.panel_canvas.bind(
            "<Leave>", lambda e: self.panel_canvas.unbind_all("<MouseWheel>"))

        # --- Camera ---
        cam_box = ttk.LabelFrame(panel, text="Camera", padding=(8, 5))
        cam_box.pack(fill=tk.X, pady=3)
        ttk.Label(cam_box, text="Index:").grid(row=0, column=0, sticky="w")
        self.cam_index_var = tk.IntVar(value=0)
        ttk.Spinbox(
            cam_box, from_=0, to=10, width=5, textvariable=self.cam_index_var
        ).grid(row=0, column=1, padx=4)
        ttk.Button(cam_box, text="Reconnect", command=self.reconnect).grid(
            row=0, column=2, padx=4
        )

        # --- Live reading ---
        read_box = ttk.LabelFrame(panel, text="Live measurement", padding=(8, 5))
        read_box.pack(fill=tk.X, pady=3)
        self.reading_var = tk.StringVar(value="-- ")
        ttk.Label(
            read_box, textvariable=self.reading_var,
            font=("Segoe UI", 22, "bold"),
        ).pack(anchor="w")
        self.detail_var = tk.StringVar(value="No fiber detected")
        ttk.Label(read_box, textvariable=self.detail_var).pack(anchor="w")

        # --- Calibration ---
        cal_box = ttk.LabelFrame(panel, text="Calibration", padding=(8, 5))
        cal_box.pack(fill=tk.X, pady=3)
        self.calib_var = tk.StringVar()
        ttk.Label(cal_box, textvariable=self.calib_var).pack(anchor="w")
        btns = ttk.Frame(cal_box)
        btns.pack(anchor="w", pady=(4, 0))
        ttk.Button(
            btns, text="Calibrate (reference)", command=self.calibrate_reference
        ).grid(row=0, column=0, padx=2)
        ttk.Button(
            btns, text="Enter factor", command=self.calibrate_manual
        ).grid(row=0, column=1, padx=2)
        ttk.Button(btns, text="Clear", command=self.clear_calibration).grid(
            row=0, column=2, padx=2
        )

        # --- Detection parameters ---
        par_box = ttk.LabelFrame(panel, text="Detection parameters", padding=(8, 5))
        par_box.pack(fill=tk.X, pady=3)
        par_box.columnconfigure(1, weight=1)

        self.otsu_var = tk.BooleanVar(value=self.detector.use_otsu)
        ttk.Checkbutton(
            par_box, text="Auto threshold (Otsu)", variable=self.otsu_var,
            command=self._sync_params,
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        ttk.Label(par_box, text="Threshold").grid(row=1, column=0, sticky="w")
        self.thresh_var = tk.IntVar(value=self.detector.manual_thresh)
        ttk.Scale(
            par_box, from_=0, to=255, variable=self.thresh_var,
            command=lambda *_: self._sync_params(), length=200,
        ).grid(row=1, column=1, sticky="we")

        ttk.Label(par_box, text="Blur").grid(row=2, column=0, sticky="w")
        self.blur_var = tk.IntVar(value=self.detector.blur_ksize)
        ttk.Scale(
            par_box, from_=1, to=21, variable=self.blur_var,
            command=lambda *_: self._sync_params(), length=200,
        ).grid(row=2, column=1, sticky="we")

        ttk.Label(par_box, text="Min area").grid(row=3, column=0, sticky="w")
        self.area_var = tk.IntVar(value=self.detector.min_area)
        ttk.Scale(
            par_box, from_=50, to=5000, variable=self.area_var,
            command=lambda *_: self._sync_params(), length=200,
        ).grid(row=3, column=1, sticky="we")
        ttk.Button(
            par_box, text="Reset to defaults", command=self.reset_parameters
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(6, 0))
        ttk.Label(
            par_box,
            text="The Processed feed (below) shows the effect of these settings.",
            wraplength=320, foreground="#555",
        ).grid(row=5, column=0, columnspan=2, sticky="w", pady=(4, 0))

        # --- Experiment ---
        exp_box = ttk.LabelFrame(panel, text="Experiment", padding=(8, 5))
        exp_box.pack(fill=tk.X, pady=3)
        exp_box.columnconfigure(1, weight=1)

        # Custom file name for the saved data.
        ttk.Label(exp_box, text="File name:").grid(row=0, column=0, sticky="w")
        self.filename_var = tk.StringVar(value="experiment")
        ttk.Entry(exp_box, textvariable=self.filename_var).grid(
            row=0, column=1, columnspan=2, sticky="we", padx=2, pady=2
        )

        # Save folder (defaults to Data/, can be redirected via the file manager).
        ttk.Label(exp_box, text="Save folder:").grid(row=1, column=0, sticky="w")
        self.save_dir_var = tk.StringVar(value=self.save_dir)
        ttk.Label(exp_box, textvariable=self.save_dir_var, foreground="#555",
                  wraplength=210).grid(row=1, column=1, sticky="w", padx=2)
        ttk.Button(exp_box, text="Change...", command=self.change_save_folder).grid(
            row=1, column=2, padx=2
        )

        # Number of samples to record. 0 = unlimited (the original behaviour);
        # any positive number auto-stops recording once that many samples are
        # captured, after which the Start / Pause & Save / Save As buttons work
        # exactly as before.
        ttk.Label(exp_box, text="Samples to record:").grid(
            row=2, column=0, sticky="w"
        )
        self.sample_target_var = tk.IntVar(value=0)
        ttk.Spinbox(
            exp_box, from_=0, to=1000000, increment=10,
            textvariable=self.sample_target_var, width=10,
        ).grid(row=2, column=1, sticky="w", padx=2, pady=2)
        ttk.Label(exp_box, text="(0 = unlimited)", foreground="#555").grid(
            row=2, column=2, sticky="w"
        )

        # Recording / saving buttons.
        self.start_btn = ttk.Button(
            exp_box, text="Start", command=self.start_recording
        )
        self.start_btn.grid(row=3, column=0, padx=2, pady=(6, 0), sticky="we")
        self.pause_btn = ttk.Button(
            exp_box, text="Pause & Save", command=self.pause_recording,
            state=tk.DISABLED,
        )
        self.pause_btn.grid(row=3, column=1, padx=2, pady=(6, 0), sticky="we")
        ttk.Button(exp_box, text="Save As...", command=self.save_as).grid(
            row=3, column=2, padx=2, pady=(6, 0), sticky="we"
        )

        self.exp_status_var = tk.StringVar(value="Idle - 0 samples")
        ttk.Label(exp_box, textvariable=self.exp_status_var).grid(
            row=4, column=0, columnspan=3, sticky="w", pady=(4, 0)
        )
        ttk.Label(
            exp_box,
            text="Saved as both .csv and .xlsx with the same name.",
            foreground="#555", wraplength=300,
        ).grid(row=5, column=0, columnspan=3, sticky="w")

        # --- Streaming to FrED Pi (over WiFi) ---
        stream_box = ttk.LabelFrame(
            panel, text="Stream to FrED Pi (WiFi)", padding=(8, 5)
        )
        stream_box.pack(fill=tk.X, pady=3)
        stream_box.columnconfigure(1, weight=1)

        # Reminder of which hotspot to join first.
        ttk.Label(
            stream_box,
            text=(f"1. Join Wi-Fi  '{PI_HOTSPOT_SSID}'  "
                  f"(password: {PI_HOTSPOT_PASSWORD})"),
            foreground="#225522", wraplength=320,
        ).grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Label(
            stream_box, text="2. Enter the Pi address shown on its screen:",
            foreground="#555", wraplength=320,
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(2, 4))

        ttk.Label(stream_box, text="Pi IP:").grid(row=2, column=0, sticky="w")
        self.host_var = tk.StringVar(value=PI_DEFAULT_IP)
        ttk.Entry(stream_box, textvariable=self.host_var, width=16).grid(
            row=2, column=1, padx=4, sticky="we"
        )

        ttk.Label(stream_box, text="Port:").grid(row=3, column=0, sticky="w")
        self.port_var = tk.StringVar(value=str(PI_DEFAULT_PORT))
        ttk.Entry(stream_box, textvariable=self.port_var, width=8).grid(
            row=3, column=1, padx=4, sticky="w"
        )

        self.connect_btn = ttk.Button(
            stream_box, text="Connect", command=self.toggle_connection
        )
        self.connect_btn.grid(row=4, column=0, pady=(6, 0), padx=2, sticky="we")
        self.stream_btn = ttk.Button(
            stream_box, text="Start streaming", command=self.toggle_streaming,
            state=tk.DISABLED,
        )
        self.stream_btn.grid(row=4, column=1, columnspan=2, pady=(6, 0), padx=2,
                             sticky="we")

        self.stream_status_var = tk.StringVar(value="Not connected")
        ttk.Label(stream_box, textvariable=self.stream_status_var,
                  wraplength=320).grid(
            row=5, column=0, columnspan=3, sticky="w", pady=(4, 0)
        )

        # --- Experiment tab (send a whole run to FrED) ---
        self._build_experiment_tab(self._experiment_tab)

        # --- Status bar ---
        self.status_var = tk.StringVar(value="Ready.")
        ttk.Label(
            self.root, textvariable=self.status_var, relief=tk.SUNKEN,
            anchor="w", padding=4,
        ).pack(fill=tk.X, side=tk.BOTTOM)

    # ------------------------------------------------------------------ #
    # Camera handling
    # ------------------------------------------------------------------ #
    def open_camera(self):
        self.camera_index = self.cam_index_var.get()
        # CAP_DSHOW avoids slow start-up on Windows.
        self.cap = cv2.VideoCapture(self.camera_index, cv2.CAP_DSHOW)
        if not self.cap.isOpened():
            self.cap = cv2.VideoCapture(self.camera_index)
        if self.cap.isOpened():
            self.status_var.set(f"Camera {self.camera_index} connected.")
        else:
            self.status_var.set(
                f"Could not open camera {self.camera_index}. "
                "Check the index and reconnect."
            )

    def reconnect(self):
        if self.recording:
            messagebox.showwarning(
                "Recording", "Stop the experiment before switching cameras."
            )
            return
        if self.cap is not None:
            self.cap.release()
        self.open_camera()

    # ------------------------------------------------------------------ #
    # Parameter syncing
    # ------------------------------------------------------------------ #
    def _sync_params(self):
        self.detector.use_otsu = self.otsu_var.get()
        self.detector.manual_thresh = int(self.thresh_var.get())
        self.detector.blur_ksize = int(self.blur_var.get())
        self.detector.min_area = int(self.area_var.get())

    def reset_parameters(self):
        """Restore the preset detection parameters."""
        d = FiberDetector.DEFAULTS
        self.otsu_var.set(d["use_otsu"])
        self.thresh_var.set(d["manual_thresh"])
        self.blur_var.set(d["blur_ksize"])
        self.area_var.set(d["min_area"])
        self._sync_params()
        self.status_var.set("Detection parameters reset to defaults.")

    # ------------------------------------------------------------------ #
    # Main loop
    # ------------------------------------------------------------------ #
    def _loop(self):
        if self.cap is not None and self.cap.isOpened():
            ok, frame = self.cap.read()
            if ok:
                annotated, mask_view, result = self.detector.process(frame)
                self.last_result = result

                self._update_reading(result)

                if self.recording and result["found"]:
                    self._record(result)

                if self.streaming:
                    self._stream(result)

                # Left = live detected frame, right = processed mask/edges.
                self._show_on(self.video_label, annotated)
                self._show_on(self.proc_label, mask_view)

        # Handle any messages the Pi sent back (experiment status / data).
        self._drain_pi_messages()
        self.root.after(15, self._loop)

    def _drain_pi_messages(self):
        """Process Pi -> laptop messages on the Tk main thread (safe for UI)."""
        while True:
            try:
                msg = self.streamer.rx_queue.get_nowait()
            except queue.Empty:
                break
            self._handle_pi_message(msg)

    def _stream(self, result):
        """Send the current measurement to the Pi, throttled to STREAM_INTERVAL."""
        now = time.time()
        if now - self._last_stream_t < self.STREAM_INTERVAL:
            return
        self._last_stream_t = now

        d_px = result["diameter_px"]
        if self.calib.is_calibrated:
            diameter = self.calib.to_real(d_px)
            units = self.calib.units
        else:
            diameter = d_px
            units = "px"
        elapsed = now - self.stream_start_time if self.stream_start_time else 0.0

        ok = self.streamer.send(diameter, units, elapsed, result["found"])
        if not ok:
            # Link dropped while streaming.
            self.streaming = False
            self.stream_btn.configure(text="Start streaming", state=tk.DISABLED)
            self.connect_btn.configure(text="Connect")
            self.stream_status_var.set("Link lost - reconnect to resume.")

    # ------------------------------------------------------------------ #
    # Display sizing / fullscreen
    # ------------------------------------------------------------------ #
    def _on_video_resize(self, event):
        """Track the video area size so each feed scales to fill it."""
        # Feeds are stacked: each uses the full width and about half the height
        # (minus the two title rows / padding).
        self._feed_w = max(160, event.width - 10)
        self._feed_h = max(120, event.height // 2 - 28)

    def _update_panel_scroll(self):
        """Size the control canvas to its content; show the scrollbar only if needed."""
        bbox = self.panel_canvas.bbox("all")
        if bbox is None:
            return
        content_w = bbox[2] - bbox[0]
        content_h = bbox[3] - bbox[1]
        # Keep the canvas exactly as wide as the controls so nothing is clipped.
        if self.panel_canvas.winfo_width() != content_w:
            self.panel_canvas.configure(width=content_w)
        self.panel_canvas.configure(scrollregion=bbox)
        if content_h > self.panel_canvas.winfo_height() + 1:
            self.panel_vsb.grid()
        else:
            self.panel_vsb.grid_remove()
            self.panel_canvas.yview_moveto(0)

    def _toggle_fullscreen(self, _event=None):
        self._fullscreen = not self._fullscreen
        self.root.attributes("-fullscreen", self._fullscreen)

    def _exit_fullscreen(self, _event=None):
        if self._fullscreen:
            self._fullscreen = False
            self.root.attributes("-fullscreen", False)

    def _show_on(self, label, bgr):
        """Scale a BGR frame to the current feed size and show it on ``label``."""
        rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        h, w = rgb.shape[:2]
        scale = min(self._feed_w / w, self._feed_h / h)
        if scale <= 0:
            scale = 1.0
        new_w, new_h = max(1, int(w * scale)), max(1, int(h * scale))
        interp = cv2.INTER_AREA if scale < 1 else cv2.INTER_LINEAR
        rgb = cv2.resize(rgb, (new_w, new_h), interpolation=interp)
        imgtk = ImageTk.PhotoImage(image=Image.fromarray(rgb))
        label.imgtk = imgtk  # keep a reference to avoid GC
        label.configure(image=imgtk)

    def _update_reading(self, result):
        if not result["found"]:
            self.reading_var.set("--")
            self.detail_var.set("No fiber detected")
            return

        d_px = result["diameter_px"]
        if self.calib.is_calibrated:
            d_real = self.calib.to_real(d_px)
            u = self.calib.units
            self.reading_var.set(f"{d_real:.4f} {u}")
            self.detail_var.set(
                f"{d_px:.1f} px | min {self.calib.to_real(result['min_px']):.4f} "
                f"max {self.calib.to_real(result['max_px']):.4f} "
                f"std {self.calib.to_real(result['std_px']):.4f} {u}"
            )
        else:
            self.reading_var.set(f"{d_px:.1f} px")
            self.detail_var.set(
                f"min {result['min_px']:.1f}  max {result['max_px']:.1f}  "
                f"std {result['std_px']:.1f} px  (not calibrated)"
            )

    # ------------------------------------------------------------------ #
    # Calibration actions
    # ------------------------------------------------------------------ #
    def _update_calib_label(self):
        if self.calib.is_calibrated:
            self.calib_var.set(
                f"Calibrated: {self.calib.units_per_pixel:.6f} "
                f"{self.calib.units}/px"
            )
        else:
            self.calib_var.set("Not calibrated (measuring in pixels)")

    def calibrate_reference(self):
        if self.last_result is None or not self.last_result["found"]:
            messagebox.showwarning(
                "Calibration",
                "No fiber detected. Place a reference object of known size in "
                "view, make sure it is detected, then calibrate.",
            )
            return
        measured_px = self.last_result["diameter_px"]
        units = simpledialog.askstring(
            "Calibration", "Units (e.g. mm, um):", initialvalue=self.calib.units,
            parent=self.root,
        )
        if not units:
            return
        known = simpledialog.askfloat(
            "Calibration",
            f"The detected object measures {measured_px:.1f} px.\n"
            f"Enter its true diameter in {units}:",
            parent=self.root, minvalue=0.0,
        )
        if known is None or known <= 0:
            return
        self.calib.set_from_reference(measured_px, known, units)
        self._update_calib_label()
        self.status_var.set(
            f"Calibrated: {self.calib.units_per_pixel:.6f} {units}/px"
        )

    def calibrate_manual(self):
        units = simpledialog.askstring(
            "Calibration", "Units (e.g. mm, um):", initialvalue=self.calib.units,
            parent=self.root,
        )
        if not units:
            return
        factor = simpledialog.askfloat(
            "Calibration", f"Enter {units} per pixel:", parent=self.root,
            minvalue=0.0,
        )
        if factor is None or factor <= 0:
            return
        self.calib.set_factor(factor, units)
        self._update_calib_label()
        self.status_var.set(f"Calibration factor set: {factor:.6f} {units}/px")

    def clear_calibration(self):
        self.calib.clear()
        self._update_calib_label()
        self.status_var.set("Calibration cleared. Measuring in pixels.")

    # ------------------------------------------------------------------ #
    # Experiment recording
    # ------------------------------------------------------------------ #
    def start_recording(self):
        if self.recording:
            return
        self.recording = True
        self.records = []
        self.frame_counter = 0
        self.exp_start_time = _dt.datetime.now()
        self.start_btn.configure(state=tk.DISABLED)
        self.pause_btn.configure(state=tk.NORMAL)
        self.status_var.set("Recording started.")

    def _record(self, result):
        now = _dt.datetime.now()
        elapsed = (now - self.exp_start_time).total_seconds()
        self.frame_counter += 1
        d_px = result["diameter_px"]
        row = {
            "timestamp": now.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            "elapsed_s": round(elapsed, 3),
            "frame": self.frame_counter,
            "diameter_px": round(d_px, 3),
            "min_px": round(result["min_px"], 3),
            "max_px": round(result["max_px"], 3),
            "std_px": round(result["std_px"], 3),
            "length_px": round(result["length_px"], 3),
            "angle_deg": round(result["angle_deg"], 3),
        }
        if self.calib.is_calibrated:
            row["diameter_real"] = round(self.calib.to_real(d_px), 6)
            row["units"] = self.calib.units
        else:
            row["diameter_real"] = ""
            row["units"] = "px"
        self.records.append(row)

        target = self._sample_target()
        n = len(self.records)
        if target and n >= target:
            self._finish_target_recording(target)
            return
        suffix = f"/{target}" if target else ""
        self.exp_status_var.set(f"Recording - {n}{suffix} samples")

    def _sample_target(self):
        """Requested number of samples (0/blank/invalid -> unlimited)."""
        try:
            target = int(self.sample_target_var.get())
        except (tk.TclError, ValueError):
            return 0
        return target if target > 0 else 0

    def _finish_target_recording(self, target):
        """Auto-stop once the requested sample count is reached.

        Recording stops and the buttons return to their normal idle state, so
        everything works just like the previous versions afterwards. The data is
        kept; the user is offered the same save as Pause & Save (and can still
        use Save As... or Start again if they decline)."""
        self.recording = False
        self.start_btn.configure(state=tk.NORMAL)
        self.pause_btn.configure(state=tk.DISABLED)
        n = len(self.records)
        self.exp_status_var.set(f"Done - {n} samples (target {target})")
        self.status_var.set(f"Reached target of {target} samples.")

        save = messagebox.askyesno(
            "Target reached",
            f"Recorded the requested {target} samples.\n\n"
            f"Save the data (.csv and .xlsx) to:\n{self.save_dir}?",
        )
        if save and self._save_dataset(self.save_dir, self._base_filename()):
            self.records = []
            self.exp_status_var.set("Idle - 0 samples")

    def change_save_folder(self):
        """Pick the folder Pause & Save writes into (Windows file manager)."""
        folder = filedialog.askdirectory(
            title="Choose folder to save experiment data",
            initialdir=self.save_dir,
        )
        if folder:
            self.save_dir = folder
            self.save_dir_var.set(folder)
            self.status_var.set(f"Save folder set to {folder}")

    def _base_filename(self):
        """Sanitised base name (no extension) from the File name box."""
        name = self.filename_var.get().strip()
        if not name:
            name = "experiment_" + _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(name)
        if ext.lower() in (".csv", ".xlsx"):
            name = base
        for ch in '<>:"/\\|?*':       # characters illegal in Windows file names
            name = name.replace(ch, "_")
        return name or "experiment"

    def pause_recording(self):
        if not self.recording:
            return
        self.recording = False
        self.start_btn.configure(state=tk.NORMAL)
        self.pause_btn.configure(state=tk.DISABLED)
        n = len(self.records)
        self.exp_status_var.set(f"Paused - {n} samples")

        if n == 0:
            messagebox.showinfo(
                "No data", "No samples were recorded; nothing to save."
            )
            return

        save = messagebox.askyesno(
            "Save experiment",
            f"Experiment paused with {n} samples.\n\n"
            f"Save the data (.csv and .xlsx) to:\n{self.save_dir}?",
        )
        if save:
            if self._save_dataset(self.save_dir, self._base_filename()):
                self.records = []
                self.exp_status_var.set("Idle - 0 samples")
        else:
            keep = messagebox.askyesno(
                "Discard or continue",
                "Data was NOT saved.\n\n"
                "Yes = discard this data\nNo = keep it (Start again to add more, "
                "Pause again to save, or use Save As...).",
            )
            if keep:
                self.records = []
                self.exp_status_var.set("Idle - 0 samples")
                self.status_var.set("Experiment data discarded.")

    def save_as(self):
        """Save the current data to a location chosen in the Windows file dialog."""
        if self.recording:
            messagebox.showwarning(
                "Recording", "Pause the experiment before saving."
            )
            return
        if not self.records:
            messagebox.showinfo("No data", "There is no recorded data to save.")
            return
        path = filedialog.asksaveasfilename(
            title="Save experiment data as",
            initialdir=self.save_dir,
            initialfile=self._base_filename() + ".csv",
            defaultextension=".csv",
            filetypes=[("CSV and Excel", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        folder = os.path.dirname(path)
        base = os.path.splitext(os.path.basename(path))[0]
        self._save_dataset(folder, base)

    def _save_dataset(self, folder, base):
        """Write <base>.csv and <base>.xlsx into ``folder``. Returns True on success."""
        try:
            os.makedirs(folder, exist_ok=True)
        except OSError as exc:
            messagebox.showerror("Save failed", f"Could not create folder:\n{exc}")
            return False

        n = len(self.records)
        csv_path = os.path.join(folder, base + ".csv")
        xlsx_path = os.path.join(folder, base + ".xlsx")

        try:
            with open(csv_path, "w", newline="", encoding="utf-8") as fh:
                writer = csv.DictWriter(fh, fieldnames=EXPORT_FIELDS)
                writer.writeheader()
                for row in self.records:
                    writer.writerow(row)
        except OSError as exc:
            messagebox.showerror("Save failed", f"Could not write CSV:\n{exc}")
            return False

        xlsx_ok, xlsx_msg = self._write_xlsx(xlsx_path)

        saved = csv_path + ("\n" + xlsx_path if xlsx_ok else "")
        note = "" if xlsx_ok else f"\n\nNote: Excel file not written ({xlsx_msg})."
        self.exp_status_var.set(f"Saved {n} samples")
        self.status_var.set(f"Saved {csv_path}")
        messagebox.showinfo(
            "Saved",
            f"Saved {n} samples to:\n{saved}\n\nDiameter: {self._summary()}{note}",
        )
        return True

    def _write_xlsx(self, path):
        """Write the records to an .xlsx file. Returns (ok, message)."""
        try:
            from openpyxl import Workbook
        except ImportError:
            return False, "openpyxl not installed (pip install openpyxl)"
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Measurements"
            ws.append(EXPORT_FIELDS)
            for row in self.records:
                ws.append([row.get(f, "") for f in EXPORT_FIELDS])
            wb.save(path)
        except Exception as exc:
            return False, str(exc)
        return True, ""

    def _summary(self):
        d = np.array([r["diameter_px"] for r in self.records], dtype=float)
        if d.size == 0:
            return "no samples"
        if self.calib.is_calibrated:
            d_real = d * self.calib.units_per_pixel
            return (f"mean {d_real.mean():.4f} {self.calib.units}, "
                    f"std {d_real.std():.4f} {self.calib.units}")
        return f"mean {d.mean():.2f} px, std {d.std():.2f} px"

    # ------------------------------------------------------------------ #
    # Streaming to the FrED Pi
    # ------------------------------------------------------------------ #
    def toggle_connection(self):
        if self.streamer.is_open:
            # Disconnect (also stops streaming).
            if self.streaming:
                self.toggle_streaming()
            self.streamer.close()
            self.connect_btn.configure(text="Connect")
            self.stream_btn.configure(state=tk.DISABLED)
            self.stream_status_var.set("Not connected")
            self.status_var.set("WiFi link closed.")
            return

        host = self.host_var.get().strip()
        port = self.port_var.get().strip()
        if not host or not port:
            messagebox.showwarning(
                "Connect", "Enter the Pi's IP address and port (shown on the "
                "Pi's screen). Default is 192.168.4.1 : 5005."
            )
            return
        try:
            port_num = int(port)
        except ValueError:
            messagebox.showwarning("Connect", f"'{port}' is not a valid port number.")
            return
        try:
            self.streamer.open(host, port_num)
        except Exception as exc:
            messagebox.showerror(
                "Connect failed",
                f"Could not connect to {host}:{port_num}\n\n{exc}\n\n"
                f"Make sure the laptop is joined to the '{PI_HOTSPOT_SSID}' "
                "Wi-Fi and the Pi program is running."
            )
            self.stream_status_var.set("Connection failed.")
            return
        self.connect_btn.configure(text="Disconnect")
        self.stream_btn.configure(state=tk.NORMAL)
        self.stream_status_var.set(f"Connected to {self.streamer.endpoint} (idle)")
        self.status_var.set(f"Connected to {self.streamer.endpoint}.")

        # Start diameter streaming immediately: once you connect, FrED is
        # already receiving diameter, so an experiment needs no stream setup.
        if not self.streaming:
            self.toggle_streaming()

    def toggle_streaming(self):
        if not self.streamer.is_open:
            return
        self.streaming = not self.streaming
        if self.streaming:
            self.stream_start_time = time.time()
            self._last_stream_t = 0.0
            self.stream_btn.configure(text="Stop streaming")
            self.stream_status_var.set(f"Streaming to {self.streamer.endpoint}...")
            self.status_var.set("Streaming diameter to the Pi.")
        else:
            self.stream_btn.configure(text="Start streaming")
            self.stream_status_var.set(f"Connected to {self.streamer.endpoint} (idle)")
            self.status_var.set("Streaming stopped.")

    # ================================================================== #
    # Experiment tab: configure a run, send it to FrED, retrieve the data
    # ================================================================== #
    def _scrollable(self, parent):
        """Create a vertically scrollable frame inside ``parent``; return it."""
        canvas = tk.Canvas(parent, borderwidth=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        inner = ttk.Frame(canvas, padding=(10, 8))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Enter>", lambda e: canvas.bind_all(
            "<MouseWheel>",
            lambda ev: canvas.yview_scroll(int(-ev.delta / 120), "units")))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        return inner

    @staticmethod
    def _exp_row(parent, label, var, row, width=10):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=2)
        ttk.Entry(parent, textvariable=var, width=width).grid(
            row=row, column=1, sticky="w", padx=4, pady=2)

    def _build_experiment_tab(self, parent):
        panel = self._scrollable(parent)

        ttk.Label(panel, text="Send a full experiment to FrED",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ttk.Label(panel, wraplength=360, foreground="#555",
                  text="Connect on the 'Measure & Stream' tab first. The "
                       "diameter is streamed from this laptop, so keep the "
                       "camera running and calibrated for diameter to be "
                       "recorded.").pack(anchor="w", pady=(0, 6))

        # --- Experiment sequence timing ---
        tbox = ttk.LabelFrame(panel, text="Experiment sequence", padding=(8, 5))
        tbox.pack(fill=tk.X, pady=3)
        self.exp_heating_delay_var = tk.StringVar(value="60")
        self.exp_heat_extrude_time_var = tk.StringVar(value="30")
        self.exp_heat_extrude_speed_var = tk.StringVar(value="1.5")
        self.exp_data_delay_var = tk.StringVar(value="10")
        self.exp_data_time_var = tk.StringVar(value="120")
        self.exp_post_spool_var = tk.StringVar(value="15")
        self._exp_row(tbox, "Heating time (s) - heater only",
                      self.exp_heating_delay_var, 0)
        self._exp_row(tbox, "Heating + extrusion time (s)",
                      self.exp_heat_extrude_time_var, 1)
        self._exp_row(tbox, "Extrusion rate during it (RPM)",
                      self.exp_heat_extrude_speed_var, 2)
        self._exp_row(tbox, "Experiment settle time (s)",
                      self.exp_data_delay_var, 3)
        self._exp_row(tbox, "Data-taking time (s)", self.exp_data_time_var, 4)
        self._exp_row(tbox, "Extra spooling after end (s)",
                      self.exp_post_spool_var, 5)
        ttk.Label(tbox, foreground="#555", wraplength=340,
                  text="Sequence: heat only -> heat + extrude (at the rate "
                       "above) -> settle with all systems on -> record -> "
                       "everything stops except the spooler, which keeps "
                       "coiling fiber for the extra spooling time.").grid(
            row=6, column=0, columnspan=2, sticky="w", pady=(4, 0))

        # --- Heater ---
        hbox = ttk.LabelFrame(panel, text="Heater", padding=(8, 5))
        hbox.pack(fill=tk.X, pady=3)
        ttk.Label(hbox, text="Mode").grid(row=0, column=0, sticky="w")
        self.exp_heater_mode_var = tk.StringVar(value="closed")
        ttk.Combobox(hbox, textvariable=self.exp_heater_mode_var, width=8,
                     state="readonly", values=["closed", "open"]).grid(
            row=0, column=1, sticky="w", padx=4)
        self.exp_target_temp_var = tk.StringVar(value="95")
        self.exp_temp_kp_var = tk.StringVar(value="1.0")
        self.exp_temp_ki_var = tk.StringVar(value="0.001")
        self.exp_temp_kd_var = tk.StringVar(value="0.05")
        self.exp_heater_pwm_var = tk.StringVar(value="0")
        self._exp_row(hbox, "Target temp (C) [closed]", self.exp_target_temp_var, 1)
        self._exp_row(hbox, "Temp Kp [closed]", self.exp_temp_kp_var, 2)
        self._exp_row(hbox, "Temp Ki [closed]", self.exp_temp_ki_var, 3)
        self._exp_row(hbox, "Temp Kd [closed]", self.exp_temp_kd_var, 4)
        self._exp_row(hbox, "Heater PWM (%) [open]", self.exp_heater_pwm_var, 5)

        # --- Spooler ---
        sbox = ttk.LabelFrame(panel, text="Spooler (DC motor)", padding=(8, 5))
        sbox.pack(fill=tk.X, pady=3)
        ttk.Label(sbox, text="Mode").grid(row=0, column=0, sticky="w")
        self.exp_spooler_mode_var = tk.StringVar(value="closed")
        ttk.Combobox(sbox, textvariable=self.exp_spooler_mode_var, width=8,
                     state="readonly", values=["closed", "open"]).grid(
            row=0, column=1, sticky="w", padx=4)
        self.exp_motor_setpoint_var = tk.StringVar(value="30")
        self.exp_motor_kp_var = tk.StringVar(value="0.5")
        self.exp_motor_ki_var = tk.StringVar(value="0.5")
        self.exp_motor_kd_var = tk.StringVar(value="0.05")
        self.exp_dc_pwm_var = tk.StringVar(value="0")
        self._exp_row(sbox, "Setpoint (RPM) [closed]", self.exp_motor_setpoint_var, 1)
        self._exp_row(sbox, "Motor Kp [closed]", self.exp_motor_kp_var, 2)
        self._exp_row(sbox, "Motor Ki [closed]", self.exp_motor_ki_var, 3)
        self._exp_row(sbox, "Motor Kd [closed]", self.exp_motor_kd_var, 4)
        self._exp_row(sbox, "DC Motor PWM (%) [open]", self.exp_dc_pwm_var, 5)

        # --- Stepper / Fan / Diameter ---
        obox = ttk.LabelFrame(panel, text="Extruder / Fan / Diameter",
                              padding=(8, 5))
        obox.pack(fill=tk.X, pady=3)
        self.exp_extrusion_var = tk.StringVar(value="1.5")
        self.exp_fan_var = tk.StringVar(value="40")
        self.exp_target_diameter_var = tk.StringVar(value="0.35")
        self._exp_row(obox, "Extrusion speed (RPM)", self.exp_extrusion_var, 0)
        self._exp_row(obox, "Fan duty (%)", self.exp_fan_var, 1)
        self._exp_row(obox, "Target diameter (mm)", self.exp_target_diameter_var, 2)

        # --- Save settings ---
        vbox = ttk.LabelFrame(panel, text="Save received data", padding=(8, 5))
        vbox.pack(fill=tk.X, pady=3)
        vbox.columnconfigure(1, weight=1)
        ttk.Label(vbox, text="File name:").grid(row=0, column=0, sticky="w")
        self.exp_name_var = tk.StringVar(value="fred_experiment")
        ttk.Entry(vbox, textvariable=self.exp_name_var).grid(
            row=0, column=1, columnspan=2, sticky="we", padx=2, pady=2)
        ttk.Label(vbox, text="Folder:").grid(row=1, column=0, sticky="w")
        self.exp_save_dir_var = tk.StringVar(value=self.exp_save_dir)
        ttk.Label(vbox, textvariable=self.exp_save_dir_var, foreground="#555",
                  wraplength=210).grid(row=1, column=1, sticky="w", padx=2)
        ttk.Button(vbox, text="Change...", command=self.change_exp_folder).grid(
            row=1, column=2, padx=2)

        # --- Action buttons ---
        btns = ttk.Frame(panel)
        btns.pack(fill=tk.X, pady=(6, 2))
        ttk.Button(btns, text="Send & Start Experiment",
                   command=self.send_experiment).pack(fill=tk.X, pady=2)
        row2 = ttk.Frame(btns)
        row2.pack(fill=tk.X)
        ttk.Button(row2, text="Abort", command=self.abort_experiment).pack(
            side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 2))
        self.exp_retrieve_btn = ttk.Button(
            row2, text="Retrieve Data", command=self.retrieve_data)
        self.exp_retrieve_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(2, 0))

        self.exp_status_var = tk.StringVar(value="Experiment: idle")
        ttk.Label(panel, textvariable=self.exp_status_var, wraplength=360,
                  font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(6, 0))

    def change_exp_folder(self):
        folder = filedialog.askdirectory(
            title="Choose folder for experiment data",
            initialdir=self.exp_save_dir)
        if folder:
            self.exp_save_dir = folder
            self.exp_save_dir_var.set(folder)

    def _exp_float(self, var, default=0.0):
        try:
            return float(var.get())
        except (tk.TclError, ValueError):
            return default

    def send_experiment(self):
        if not self.streamer.is_open:
            messagebox.showwarning(
                "Not connected", "Connect to FrED on the 'Measure & Stream' "
                "tab first.")
            return
        try:
            heating = float(self.exp_heating_delay_var.get())
            heat_extrude = float(self.exp_heat_extrude_time_var.get())
            heat_extrude_rpm = float(self.exp_heat_extrude_speed_var.get())
            data_delay = float(self.exp_data_delay_var.get())
            data_time = float(self.exp_data_time_var.get())
            post_spool = float(self.exp_post_spool_var.get())
        except ValueError:
            messagebox.showwarning("Timing", "Every sequence field must be a "
                                   "number (times in seconds, rate in RPM).")
            return
        if data_time <= 0:
            messagebox.showwarning("Timing", "Data-taking time must be > 0.")
            return
        if min(heating, heat_extrude, data_delay, post_spool,
               heat_extrude_rpm) < 0:
            messagebox.showwarning("Timing", "Sequence values cannot be "
                                   "negative.")
            return

        params = {
            "name": self.exp_name_var.get().strip() or "fred_experiment",
            "heater_mode": self.exp_heater_mode_var.get(),
            "target_temperature": self._exp_float(self.exp_target_temp_var, 0),
            "temp_kp": self._exp_float(self.exp_temp_kp_var, 0),
            "temp_ki": self._exp_float(self.exp_temp_ki_var, 0),
            "temp_kd": self._exp_float(self.exp_temp_kd_var, 0),
            "heater_pwm": self._exp_float(self.exp_heater_pwm_var, 0),
            "spooler_mode": self.exp_spooler_mode_var.get(),
            "motor_setpoint": self._exp_float(self.exp_motor_setpoint_var, 0),
            "motor_kp": self._exp_float(self.exp_motor_kp_var, 0),
            "motor_ki": self._exp_float(self.exp_motor_ki_var, 0),
            "motor_kd": self._exp_float(self.exp_motor_kd_var, 0),
            "dc_motor_pwm": self._exp_float(self.exp_dc_pwm_var, 0),
            "extrusion_speed": self._exp_float(self.exp_extrusion_var, 0),
            "fan_duty": self._exp_float(self.exp_fan_var, 0),
            "target_diameter": self._exp_float(self.exp_target_diameter_var, 0.35),
            "heating_delay": heating,
            "heat_extrude_time": heat_extrude,
            "heat_extrude_speed": heat_extrude_rpm,
            "data_delay": data_delay,
            "data_taking_time": data_time,
            "post_spool_time": post_spool,
        }

        if not self.calib.is_calibrated:
            if not messagebox.askyesno(
                "Not calibrated",
                "The camera is not calibrated, so diameter will be recorded in "
                "pixels, not mm. Send the experiment anyway?"):
                return

        # Make sure diameter is actually streaming so FrED can record it.
        if not self.streaming:
            self.stream_start_time = time.time()
            self._last_stream_t = 0.0
            self.streaming = True
            self.stream_btn.configure(text="Stop streaming")
            self.stream_status_var.set(f"Streaming to {self.streamer.endpoint}...")

        if self.streamer.send_command({"type": "experiment", "params": params}):
            self.exp_data_ready = False
            total = heating + heat_extrude + data_delay + data_time
            spool_note = f" + {post_spool:.0f}s spooling" if post_spool else ""
            self.exp_status_var.set(
                f"Experiment sent. Heating... (total ~{total:.0f}s{spool_note})")
            self.status_var.set("Experiment sent to FrED.")
        else:
            messagebox.showerror("Send failed",
                                 "Could not send the experiment (link lost).")

    def abort_experiment(self):
        if not self.streamer.is_open:
            return
        if messagebox.askyesno("Abort", "Abort the running experiment on FrED?"):
            self.streamer.send_command({"type": "abort"})
            self.exp_status_var.set("Abort sent.")

    def retrieve_data(self):
        if not self.streamer.is_open:
            messagebox.showwarning("Not connected", "Connect to FrED first.")
            return
        if self.streamer.send_command({"type": "get_data"}):
            self.status_var.set("Requested experiment data from FrED...")

    # ------------------------------------------------------------------ #
    # Messages coming back from FrED (run on the Tk main thread)
    # ------------------------------------------------------------------ #
    def _handle_pi_message(self, msg):
        mtype = msg.get("type")
        if mtype == "status":
            phase = msg.get("phase", "")
            text = msg.get("message", phase)
            remaining = msg.get("remaining", 0)
            extra = f" ({remaining:.0f}s left)" if remaining else ""
            self.exp_status_var.set(f"Experiment: {text}{extra}")
            if msg.get("data_ready") or phase == "complete":
                self.exp_data_ready = True
                self.exp_status_var.set(
                    f"Experiment: {text} - click 'Retrieve Data'.")
        elif mtype == "event" and msg.get("event") == "no_data":
            messagebox.showinfo(
                "No data", msg.get("message", "No experiment data available "
                "yet. Run an experiment first."))
        elif mtype == "data":
            self._receive_experiment_data(msg)

    def _receive_experiment_data(self, msg):
        try:
            csv_text = base64.b64decode(msg.get("b64", "")).decode("utf-8")
        except Exception as exc:
            messagebox.showerror("Receive failed",
                                 f"Could not decode the data: {exc}")
            return
        base = msg.get("name") or self.exp_name_var.get().strip() or "fred_experiment"
        for ch in '<>:"/\\|?*':
            base = base.replace(ch, "_")
        self._save_experiment_data(base, csv_text)

    def _save_experiment_data(self, base, csv_text):
        folder = self.exp_save_dir
        try:
            os.makedirs(folder, exist_ok=True)
        except OSError as exc:
            messagebox.showerror("Save failed", f"Could not create folder:\n{exc}")
            return
        csv_path = os.path.join(folder, base + ".csv")
        xlsx_path = os.path.join(folder, base + ".xlsx")
        try:
            with open(csv_path, "w", newline="", encoding="utf-8") as fh:
                fh.write(csv_text)
        except OSError as exc:
            messagebox.showerror("Save failed", f"Could not write CSV:\n{exc}")
            return
        xlsx_ok, xlsx_msg = self._exp_write_xlsx(xlsx_path, csv_text)
        saved = csv_path + ("\n" + xlsx_path if xlsx_ok else "")
        note = "" if xlsx_ok else f"\n\nNote: Excel not written ({xlsx_msg})."
        self.exp_status_var.set("Experiment data saved.")
        self.status_var.set(f"Saved {csv_path}")
        messagebox.showinfo("Experiment data saved",
                            f"Saved FrED experiment data to:\n{saved}{note}")

    @staticmethod
    def _exp_write_xlsx(path, csv_text):
        """Write the received wide table into a formatted .xlsx.

        FrED sends a semicolon-delimited CSV with comma decimals (for Excel in
        es-MX). Numeric cells are converted back to real numbers, the header
        row is bold white on a color per subsystem, and three native Excel
        line charts (Diameter, Temperature, Spooler RPM vs time - the same
        graphs the Pi shows on screen) are placed right next to the data.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.chart import Reference, ScatterChart, Series
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return False, "openpyxl not installed (pip install openpyxl)"
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "FrED Experiment"
            reader = csv.reader(io.StringIO(csv_text), delimiter=";")
            for row in reader:
                out = []
                for cell in row:
                    c = cell.strip()
                    if c == "":
                        out.append("")
                        continue
                    try:
                        out.append(float(c.replace(",", ".")))
                    except ValueError:
                        out.append(cell)
                ws.append(out)

            headers = [str(c.value) if c.value is not None else ""
                       for c in ws[1]]
            n_rows = ws.max_row

            # ---- header row: bold white text on a color per subsystem ---- #
            def header_color(name):
                n = name.lower()
                if n.startswith("time"):
                    return "595959"        # gray
                if n.startswith("temp"):
                    return "C0504D"        # red
                if n.startswith("diameter"):
                    return "4472C4"        # blue
                if n.startswith("fan"):
                    return "31859C"        # teal
                if n.startswith("extruder"):
                    return "7030A0"        # purple
                if n.startswith("spooler"):
                    return "548235"        # green
                return "444444"

            for idx, name in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=header_color(name))
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(idx)].width = \
                    max(11, min(20, len(name) + 2))
            ws.row_dimensions[1].height = 30
            ws.freeze_panes = "A2"

            # ---- the Pi's three graphs as native Excel charts ------------ #
            if n_rows > 2 and "Time (s)" in headers:
                t_col = headers.index("Time (s)") + 1
                xref = Reference(ws, min_col=t_col, min_row=2, max_row=n_rows)

                def make_chart(title, y_title, series_names):
                    chart = ScatterChart()
                    chart.title = title
                    chart.style = 13
                    chart.x_axis.title = "Time (s)"
                    chart.y_axis.title = y_title
                    chart.x_axis.delete = False
                    chart.y_axis.delete = False
                    chart.height = 9
                    chart.width = 20
                    for name in series_names:
                        if name not in headers:
                            continue
                        col = headers.index(name) + 1
                        yref = Reference(ws, min_col=col, min_row=1,
                                         max_row=n_rows)
                        series = Series(yref, xref, title_from_data=True)
                        series.marker.symbol = "none"
                        series.smooth = False
                        chart.series.append(series)
                    return chart if chart.series else None

                specs = [
                    ("Diameter", "Diameter (mm)",
                     ["Diameter (mm)", "Diameter raw (mm)",
                      "Diameter setpoint (mm)"]),
                    ("Temperature", "Temperature (C)",
                     ["Temperature (C)", "Temp setpoint (C)"]),
                    ("DC Spooling Motor", "Speed (RPM)",
                     ["Spooler RPM", "Spooler setpoint (RPM)"]),
                ]
                anchor_col = get_column_letter(len(headers) + 2)
                anchor_row = 2
                for title, y_title, names in specs:
                    chart = make_chart(title, y_title, names)
                    if chart is not None:
                        ws.add_chart(chart, f"{anchor_col}{anchor_row}")
                        anchor_row += 19   # stack the charts vertically

            wb.save(path)
        except Exception as exc:
            return False, str(exc)
        return True, ""

    # ------------------------------------------------------------------ #
    # Shutdown
    # ------------------------------------------------------------------ #
    def on_close(self):
        if self.recording and self.records:
            if not messagebox.askyesno(
                "Quit",
                "An experiment is still recording and has unsaved data.\n"
                "Quit anyway? (unsaved data will be lost)",
            ):
                return
        self.streamer.close()
        if self.cap is not None:
            self.cap.release()
        self.root.destroy()


def main():
    root = tk.Tk()
    FiberApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
